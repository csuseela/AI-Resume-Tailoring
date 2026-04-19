from __future__ import annotations

import logging
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List

from fastapi import APIRouter, HTTPException, UploadFile, File
from fastapi.responses import FileResponse, HTMLResponse

logger = logging.getLogger(__name__)
router = APIRouter()

_container: Dict[str, Any] = {}


def set_container(container: Dict[str, Any]) -> None:
    global _container
    _container = container


@router.get("/", response_class=HTMLResponse)
def dashboard() -> HTMLResponse:
    html_path = Path(__file__).resolve().parent.parent / "static" / "dashboard.html"
    if not html_path.exists():
        raise HTTPException(status_code=404, detail="Dashboard not found")
    return HTMLResponse(content=html_path.read_text(encoding="utf-8"))


@router.get("/health")
def health_check() -> Dict[str, str]:
    return {"status": "healthy"}


@router.post("/workflow/run")
def run_workflow() -> Dict[str, Any]:
    wf = _container.get("workflow_service")
    if not wf:
        raise HTTPException(status_code=500, detail="Workflow service not initialized")
    result = wf.run_daily_workflow()
    return result


@router.get("/workflow/history")
def workflow_history() -> Any:
    tracker = _container.get("tracker_service")
    if not tracker:
        raise HTTPException(status_code=500, detail="Tracker not initialized")
    runs = tracker.get_history(limit=20)
    return [
        {
            "id": r.id,
            "run_date": str(r.run_date),
            "status": r.status,
            "jobs_found": r.jobs_found,
            "jobs_processed": r.jobs_processed,
        }
        for r in runs
    ]


@router.get("/api/tracker")
def get_tracker_data() -> List[Dict[str, Any]]:
    settings = _container.get("settings")
    if not settings:
        return []

    xlsx_path = settings.tracker_xlsx_path
    if not xlsx_path.exists():
        return []

    try:
        from openpyxl import load_workbook

        wb = load_workbook(str(xlsx_path))
        ws = wb.active
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        rows: List[Dict[str, Any]] = []
        key_map = {
            "Run Date": "run_date", "Company": "company", "Role": "role",
            "Location": "location", "Remote Type": "remote_type", "Posted": "posted",
            "Match %": "relevance", "ATS %": "ats_score", "H1B": "h1b_sponsor",
            "Reason": "reason", "Apply Link": "apply_link", "Tailored Resume": "output_path",
            "Status": "status",
        }

        for row in ws.iter_rows(min_row=2, values_only=False):
            entry: Dict[str, Any] = {}
            for idx, cell in enumerate(row):
                header = headers[idx] if idx < len(headers) else f"col_{idx}"
                key = key_map.get(header, header)
                value = cell.value or ""
                if key == "apply_link" and cell.hyperlink:
                    value = cell.hyperlink.target or value
                    key = "apply_url"
                elif key == "apply_link":
                    key = "apply_url"
                entry[key] = str(value)
            rows.append(entry)
        wb.close()
        return rows
    except Exception as exc:
        logger.error("Failed to read tracker: %s", exc)
        return []


@router.get("/api/download/{filename}")
def download_file(filename: str) -> FileResponse:
    settings = _container.get("settings")
    if not settings:
        raise HTTPException(status_code=500, detail="Settings not initialized")

    file_path = settings.output_dir / filename
    if not file_path.exists() or not file_path.is_file():
        raise HTTPException(status_code=404, detail="File not found")

    if ".." in filename or "/" in filename:
        raise HTTPException(status_code=400, detail="Invalid filename")

    media_type = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
    if filename.endswith(".md"):
        media_type = "text/markdown"
    elif filename.endswith(".html"):
        media_type = "text/html"
    elif filename.endswith(".xlsx"):
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    return FileResponse(path=str(file_path), filename=filename, media_type=media_type)


@router.get("/api/download-daily-excel")
def download_daily_excel(date: str = "") -> FileResponse:
    """Download the date-stamped daily Excel file (YYYY-MM-DD)."""
    settings = _container.get("settings")
    if not settings:
        raise HTTPException(status_code=500, detail="Settings not initialized")

    if not date:
        from datetime import datetime
        date = datetime.now().strftime("%Y-%m-%d")

    daily_path = settings.output_dir / f"job_tracker_{date}.xlsx"
    if not daily_path.exists():
        raise HTTPException(
            status_code=404,
            detail=f"No Excel file found for {date}. Run the workflow first.",
        )

    return FileResponse(
        path=str(daily_path),
        filename=f"job_tracker_{date}.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@router.get("/api/email-preview")
def get_latest_email_preview() -> HTMLResponse:
    settings = _container.get("settings")
    if not settings:
        raise HTTPException(status_code=500, detail="Settings not initialized")

    previews = sorted(settings.output_dir.glob("email_preview_*.html"), reverse=True)
    if not previews:
        raise HTTPException(status_code=404, detail="No email preview found")

    return HTMLResponse(content=previews[0].read_text(encoding="utf-8"))


ALLOWED_RESUME_EXTENSIONS = {".md", ".txt", ".docx", ".pdf"}
MAX_RESUME_SIZE_MB = 10


@router.get("/api/resume-info")
def get_resume_info() -> Dict[str, Any]:
    settings = _container.get("settings")
    if not settings:
        raise HTTPException(status_code=500, detail="Settings not initialized")

    resume_dir = settings.resume_dir
    if not resume_dir.exists():
        return {"has_resume": False, "files": []}

    files = []
    for ext in ALLOWED_RESUME_EXTENSIONS:
        for f in resume_dir.glob(f"*{ext}"):
            stat = f.stat()
            files.append({
                "name": f.name,
                "size": stat.st_size,
                "modified": datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M"),
            })

    files.sort(key=lambda x: x["name"])
    return {"has_resume": len(files) > 0, "files": files}


@router.post("/api/upload-resume")
async def upload_resume(file: UploadFile = File(...)) -> Dict[str, Any]:
    settings = _container.get("settings")
    if not settings:
        raise HTTPException(status_code=500, detail="Settings not initialized")

    if not file.filename:
        raise HTTPException(status_code=400, detail="No filename provided")

    ext = Path(file.filename).suffix.lower()
    if ext not in ALLOWED_RESUME_EXTENSIONS:
        raise HTTPException(
            status_code=400,
            detail=f"Unsupported file type '{ext}'. Allowed: {', '.join(sorted(ALLOWED_RESUME_EXTENSIONS))}",
        )

    content = await file.read()
    size_mb = len(content) / (1024 * 1024)
    if size_mb > MAX_RESUME_SIZE_MB:
        raise HTTPException(
            status_code=400,
            detail=f"File too large ({size_mb:.1f} MB). Max allowed: {MAX_RESUME_SIZE_MB} MB",
        )

    resume_dir = settings.resume_dir
    resume_dir.mkdir(parents=True, exist_ok=True)

    safe_name = Path(file.filename).name
    if ".." in safe_name or "/" in safe_name:
        raise HTTPException(status_code=400, detail="Invalid filename")

    dest = resume_dir / safe_name
    dest.write_bytes(content)
    logger.info("Resume uploaded: %s (%d bytes)", safe_name, len(content))

    return {
        "message": f"Resume '{safe_name}' uploaded successfully",
        "filename": safe_name,
        "size": len(content),
    }


@router.delete("/api/resume/{filename}")
def delete_resume(filename: str) -> Dict[str, str]:
    settings = _container.get("settings")
    if not settings:
        raise HTTPException(status_code=500, detail="Settings not initialized")

    if ".." in filename or "/" in filename:
        raise HTTPException(status_code=400, detail="Invalid filename")

    file_path = settings.resume_dir / filename
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="Resume not found")

    file_path.unlink()
    logger.info("Resume deleted: %s", filename)
    return {"message": f"Resume '{filename}' deleted successfully"}
