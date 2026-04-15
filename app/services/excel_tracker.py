from __future__ import annotations

import logging
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

HEADERS = [
    "Run Date", "Company", "Role", "Location", "Remote Type", "Posted",
    "Match %", "ATS %", "H1B", "Reason", "Apply Link", "Tailored Resume", "Status",
]

COL_WIDTHS = [14, 18, 26, 22, 12, 12, 10, 10, 8, 50, 14, 40, 10]

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
LINK_FONT = Font(color="1D4ED8", underline="single")


class ExcelTrackerService:
    def __init__(self, xlsx_path: Path) -> None:
        self.xlsx_path = xlsx_path
        self.xlsx_path.parent.mkdir(parents=True, exist_ok=True)

    def append_rows(self, rows: List[Dict[str, Any]]) -> Path:
        if self.xlsx_path.exists():
            wb = load_workbook(str(self.xlsx_path))
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Job Tracker"
            ws.append(HEADERS)
            for idx, width in enumerate(COL_WIDTHS, 1):
                ws.column_dimensions[get_column_letter(idx)].width = width
            for cell in ws[1]:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.alignment = Alignment(horizontal="center")

        run_date = datetime.now().strftime("%Y-%m-%d %H:%M")
        for row in rows:
            apply_url = row.get("apply_url", "")
            resume_path = row.get("output_path", "")

            values = [
                run_date, row.get("company", ""), row.get("role", ""),
                row.get("location", ""), row.get("remote_type", ""), row.get("posted", ""),
                row.get("relevance", ""), row.get("ats_score", ""),
                row.get("h1b_sponsor", ""),
                row.get("reason", ""), apply_url, resume_path, row.get("status", ""),
            ]
            ws.append(values)

            current_row = ws.max_row
            if apply_url:
                cell = ws.cell(row=current_row, column=11)
                cell.hyperlink = apply_url
                cell.font = LINK_FONT
                cell.value = "Apply"
            if resume_path:
                cell = ws.cell(row=current_row, column=12)
                cell.hyperlink = resume_path
                cell.font = LINK_FONT
                cell.value = resume_path

        wb.save(str(self.xlsx_path))
        logger.info("Excel tracker updated: %s (%d rows added)", self.xlsx_path, len(rows))
        return self.xlsx_path
